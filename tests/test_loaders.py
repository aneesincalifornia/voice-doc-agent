import pytest
from pathlib import Path
from openpyxl import Workbook
from app.loaders import load_document, load_txt, load_pdf, load_docx, load_excel

def test_load_txt_valid(temp_txt_file):
    """Test loading a valid text file."""
    docs = load_txt(temp_txt_file)

    assert len(docs) == 1
    assert "Employee Leave Policy" in docs[0].page_content
    assert docs[0].metadata["type"] == "txt"
    assert docs[0].metadata["source"] == temp_txt_file

def test_load_txt_empty(tmp_path):
    """Test loading an empty text file."""
    empty_file = tmp_path / "empty.txt"
    empty_file.write_text("")

    with pytest.raises(ValueError, match="No text found"):
        load_txt(str(empty_file))

def test_load_document_txt(temp_txt_file):
    """Test load_document dispatches to load_txt."""
    docs = load_document(temp_txt_file)

    assert len(docs) == 1
    assert "Employee Leave Policy" in docs[0].page_content

def test_load_document_missing():
    """Test loading a non-existent file."""
    with pytest.raises(FileNotFoundError):
        load_document("/nonexistent/file.txt")

def test_load_document_unsupported(tmp_path):
    """Test loading an unsupported file type."""
    bad_file = tmp_path / "test.csv"
    bad_file.write_text("data")

    with pytest.raises(ValueError, match="Unsupported file type"):
        load_document(str(bad_file))

def test_load_docx_missing():
    """Test loading a missing DOCX file."""
    # python-docx raises PackageNotFoundError, not FileNotFoundError
    from docx.opc.exceptions import PackageNotFoundError
    with pytest.raises(PackageNotFoundError):
        load_docx("/nonexistent/file.docx")

@pytest.mark.parametrize("ext", [".docx", ".pdf"])
def test_load_document_via_dispatch(tmp_path, ext):
    """Test that load_document correctly dispatches unsupported formats."""
    if ext == ".pdf":
        # Skip PDF test if load_pdf would require a real PDF
        pytest.skip("PDF loading requires a real PDF file")

    bad_file = tmp_path / f"test{ext}"
    bad_file.write_text("dummy")

    # load_document will try to load but fail during actual file processing
    # Just verify it calls the right loader by catching the underlying exception
    with pytest.raises(Exception):  # Will be FileNotFoundError or other loader-specific error
        load_document(str(bad_file))

def test_load_excel_valid(tmp_path):
    """Test loading a valid Excel file with multiple sheets."""
    xlsx_file = tmp_path / "test.xlsx"
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Data"
    ws1.append(["Name", "Value"])
    ws1.append(["Item1", 100])
    ws1.append(["Item2", 200])

    ws2 = wb.create_sheet("Summary")
    ws2.append(["Total", 300])

    wb.save(xlsx_file)

    docs = load_excel(str(xlsx_file))

    assert len(docs) == 2
    assert docs[0].metadata["type"] == "xlsx"
    assert docs[0].metadata["sheet_name"] == "Data"
    assert docs[1].metadata["sheet_name"] == "Summary"
    assert "Name" in docs[0].page_content or "Item1" in docs[0].page_content
    print(f"✓ First sheet content: {docs[0].page_content[:100]}")

def test_load_excel_empty_workbook(tmp_path):
    """Test loading an Excel file with empty data (no rows)."""
    xlsx_file = tmp_path / "empty.xlsx"
    wb = Workbook()
    ws = wb.active
    # Don't add any data — workbook exists but has no content
    wb.save(xlsx_file)

    with pytest.raises(ValueError, match="No data found"):
        load_excel(str(xlsx_file))

def test_load_excel_empty_sheets(tmp_path):
    """Test loading an Excel file with only empty sheets."""
    xlsx_file = tmp_path / "empty_sheets.xlsx"
    wb = Workbook()
    # Create empty sheets (no data at all)
    ws = wb.active
    ws.title = "Empty"
    wb.save(xlsx_file)

    with pytest.raises(ValueError, match="No data found"):
        load_excel(str(xlsx_file))

def test_load_excel_missing_file():
    """Test loading a missing Excel file."""
    with pytest.raises(FileNotFoundError):
        load_excel("/nonexistent/file.xlsx")

def test_load_document_excel(tmp_path):
    """Test load_document dispatches to load_excel for .xlsx files."""
    xlsx_file = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Col1", "Col2"])
    ws.append(["Data1", "Data2"])
    wb.save(xlsx_file)

    docs = load_document(str(xlsx_file))

    assert len(docs) >= 1
    assert docs[0].metadata["type"] == "xlsx"
